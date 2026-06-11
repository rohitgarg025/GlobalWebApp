"""Business logic for the Quantity Sheet module."""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import and_
from sqlalchemy.orm import Session

from app.db.models import (
    QsProject, QsActivity, QsFloor,
    QsBaseline, QsMonthlyEntry, QsBaselineChangeLog,
)


# ─── Projects ────────────────────────────────────────────────────────────────

def list_projects(db: Session) -> list[dict]:
    rows = db.query(QsProject).order_by(QsProject.name).all()
    return [_project_to_dict(r) for r in rows]


def create_project(db: Session, name: str) -> dict:
    if db.query(QsProject).filter_by(name=name).first():
        raise HTTPException(400, f"Project '{name}' already exists")
    proj = QsProject(name=name)
    db.add(proj)
    db.commit()
    db.refresh(proj)
    return _project_to_dict(proj)


def delete_project(db: Session, project_id: int) -> None:
    proj = _get_project(db, project_id)
    db.delete(proj)
    db.commit()


def _project_to_dict(p: QsProject) -> dict:
    return {"id": p.id, "name": p.name, "created_at": p.created_at.isoformat()}


def _get_project(db: Session, project_id: int) -> QsProject:
    proj = db.query(QsProject).filter_by(id=project_id).first()
    if not proj:
        raise HTTPException(404, "Project not found")
    return proj


# ─── Floors ──────────────────────────────────────────────────────────────────

def list_floors(db: Session, project_id: int) -> list[dict]:
    _get_project(db, project_id)
    rows = (
        db.query(QsFloor)
        .filter_by(project_id=project_id)
        .order_by(QsFloor.display_order, QsFloor.name)
        .all()
    )
    return [_floor_to_dict(r) for r in rows]


def create_floor(db: Session, project_id: int, name: str, display_order: int) -> dict:
    _get_project(db, project_id)
    existing = db.query(QsFloor).filter_by(project_id=project_id, name=name).first()
    if existing:
        raise HTTPException(400, f"Floor '{name}' already exists in this project")
    floor = QsFloor(project_id=project_id, name=name, display_order=display_order)
    db.add(floor)
    db.commit()
    db.refresh(floor)
    return _floor_to_dict(floor)


def delete_floor(db: Session, floor_id: int) -> None:
    floor = db.query(QsFloor).filter_by(id=floor_id).first()
    if not floor:
        raise HTTPException(404, "Floor not found")
    db.delete(floor)
    db.commit()


def reorder_floors(db: Session, project_id: int, floor_ids: list[int]) -> list[dict]:
    """Reorder floors by providing an ordered list of floor IDs."""
    _get_project(db, project_id)
    for order, fid in enumerate(floor_ids):
        floor = db.query(QsFloor).filter_by(id=fid, project_id=project_id).first()
        if floor:
            floor.display_order = order
    db.commit()
    return list_floors(db, project_id)


def _floor_to_dict(f: QsFloor) -> dict:
    return {"id": f.id, "project_id": f.project_id, "name": f.name, "display_order": f.display_order}


# ─── Activities ───────────────────────────────────────────────────────────────

def list_activities(db: Session) -> list[dict]:
    rows = db.query(QsActivity).order_by(QsActivity.name).all()
    return [_activity_to_dict(r) for r in rows]


def create_activity(db: Session, name: str, unit: str) -> dict:
    if db.query(QsActivity).filter_by(name=name).first():
        raise HTTPException(400, f"Activity '{name}' already exists")
    act = QsActivity(name=name, unit=unit)
    db.add(act)
    db.commit()
    db.refresh(act)
    return _activity_to_dict(act)


def delete_activity(db: Session, activity_id: int) -> None:
    act = db.query(QsActivity).filter_by(id=activity_id).first()
    if not act:
        raise HTTPException(404, "Activity not found")
    db.delete(act)
    db.commit()


def _activity_to_dict(a: QsActivity) -> dict:
    return {"id": a.id, "name": a.name, "unit": a.unit}


# ─── Data (baselines + monthly entries) ──────────────────────────────────────

def get_project_data(db: Session, project_id: int, month: str) -> dict:
    """Return combined baseline + monthly entry data for a project/month."""
    proj = _get_project(db, project_id)
    floors = list_floors(db, project_id)
    activities = list_activities(db)

    floor_ids = [f["id"] for f in floors]
    activity_ids = [a["id"] for a in activities]

    baselines: dict[tuple, QsBaseline] = {}
    if floor_ids and activity_ids:
        for b in db.query(QsBaseline).filter(
            QsBaseline.project_id == project_id,
            QsBaseline.floor_id.in_(floor_ids),
            QsBaseline.activity_id.in_(activity_ids),
        ).all():
            baselines[(b.activity_id, b.floor_id)] = b

    entries: dict[tuple, QsMonthlyEntry] = {}
    if floor_ids and activity_ids:
        for e in db.query(QsMonthlyEntry).filter(
            QsMonthlyEntry.project_id == project_id,
            QsMonthlyEntry.month == month,
            QsMonthlyEntry.floor_id.in_(floor_ids),
            QsMonthlyEntry.activity_id.in_(activity_ids),
        ).all():
            entries[(e.activity_id, e.floor_id)] = e

    cells = []
    for act in activities:
        for floor in floors:
            key = (act["id"], floor["id"])
            baseline = baselines.get(key)
            entry = entries.get(key)
            cells.append({
                "activity_id": act["id"],
                "floor_id": floor["id"],
                "total_estimated_qty": baseline.total_estimated_qty if baseline else None,
                "baseline_locked": baseline is not None,
                "estimate_actual_qty_till_date": entry.estimate_actual_qty_till_date if entry else None,
                "actual_qty": entry.actual_qty if entry else None,
                "justification": entry.justification if entry else None,
            })

    return {
        "project": _project_to_dict(proj),
        "floors": floors,
        "activities": activities,
        "month": month,
        "cells": cells,
    }


def submit_entries(
    db: Session,
    project_id: int,
    month: str,
    submitted_by: str,
    rows: list[dict],
) -> dict:
    """Validate and upsert a batch of quantity entries for a project/month."""
    _get_project(db, project_id)

    errors: list[str] = []

    for row in rows:
        act_id = row["activity_id"]
        floor_id = row["floor_id"]
        total_est = row.get("total_estimated_qty")
        est_actual = row.get("estimate_actual_qty_till_date")
        actual = row.get("actual_qty")
        justification = (row.get("justification") or "").strip()
        change_reason = (row.get("baseline_change_reason") or "").strip()

        act = db.query(QsActivity).filter_by(id=act_id).first()
        floor = db.query(QsFloor).filter_by(id=floor_id).first()
        label = f"Activity {act_id} / Floor {floor_id}"
        if act and floor:
            label = f"'{act.name}' / '{floor.name}'"

        if est_actual is None:
            errors.append(f"{label}: Estimate of Actual Qty Till Date is required")
        if actual is None:
            errors.append(f"{label}: Actual Qty is required")

        if total_est is None:
            errors.append(f"{label}: Total Estimated Qty is required")
            continue

        # Overrun check against what the baseline will be
        baseline = db.query(QsBaseline).filter_by(
            project_id=project_id, activity_id=act_id, floor_id=floor_id
        ).first()
        effective_estimate = baseline.total_estimated_qty if baseline else total_est
        if actual is not None and actual > effective_estimate and not justification:
            errors.append(
                f"{label}: Actual Qty ({actual}) exceeds Total Estimated Qty ({effective_estimate}) — justification required"
            )

        # Baseline change check
        if baseline and abs(baseline.total_estimated_qty - total_est) > 1e-9:
            if not change_reason:
                errors.append(
                    f"{label}: Total Estimated Qty changed from {baseline.total_estimated_qty} to {total_est} — reason required"
                )

    if errors:
        raise HTTPException(422, detail={"errors": errors})

    # Apply changes
    now = datetime.utcnow()
    for row in rows:
        act_id = row["activity_id"]
        floor_id = row["floor_id"]
        total_est = row["total_estimated_qty"]
        est_actual = row["estimate_actual_qty_till_date"]
        actual = row["actual_qty"]
        justification = (row.get("justification") or "").strip() or None
        change_reason = (row.get("baseline_change_reason") or "").strip()

        baseline = db.query(QsBaseline).filter_by(
            project_id=project_id, activity_id=act_id, floor_id=floor_id
        ).first()

        if baseline is None:
            baseline = QsBaseline(
                project_id=project_id,
                activity_id=act_id,
                floor_id=floor_id,
                total_estimated_qty=total_est,
                locked_by=submitted_by,
                locked_at=now,
            )
            db.add(baseline)
            db.flush()
        elif abs(baseline.total_estimated_qty - total_est) > 1e-9:
            log = QsBaselineChangeLog(
                baseline_id=baseline.id,
                old_value=baseline.total_estimated_qty,
                new_value=total_est,
                reason=change_reason,
                changed_by=submitted_by,
                changed_at=now,
            )
            db.add(log)
            baseline.total_estimated_qty = total_est

        entry = db.query(QsMonthlyEntry).filter_by(
            project_id=project_id,
            activity_id=act_id,
            floor_id=floor_id,
            month=month,
        ).first()

        if entry is None:
            entry = QsMonthlyEntry(
                project_id=project_id,
                activity_id=act_id,
                floor_id=floor_id,
                month=month,
            )
            db.add(entry)

        entry.estimate_actual_qty_till_date = est_actual
        entry.actual_qty = actual
        entry.justification = justification
        entry.submitted_by = submitted_by
        entry.submitted_at = now

    db.commit()
    return {"status": "ok", "submitted": len(rows)}


# ─── Overruns ────────────────────────────────────────────────────────────────

def get_overruns(db: Session, project_id: int | None = None) -> list[dict]:
    """Return all monthly entries where actual_qty exceeds the baseline."""
    baselines = db.query(QsBaseline)
    if project_id:
        baselines = baselines.filter_by(project_id=project_id)
    baselines = baselines.all()

    result = []
    for b in baselines:
        entries = db.query(QsMonthlyEntry).filter_by(
            project_id=b.project_id,
            activity_id=b.activity_id,
            floor_id=b.floor_id,
        ).all()
        for e in entries:
            if e.actual_qty > b.total_estimated_qty:
                proj = db.query(QsProject).filter_by(id=b.project_id).first()
                act = db.query(QsActivity).filter_by(id=b.activity_id).first()
                floor = db.query(QsFloor).filter_by(id=b.floor_id).first()
                result.append({
                    "entry_id": e.id,
                    "project": proj.name if proj else b.project_id,
                    "activity": act.name if act else b.activity_id,
                    "unit": act.unit if act else "",
                    "floor": floor.name if floor else b.floor_id,
                    "month": e.month,
                    "total_estimated_qty": b.total_estimated_qty,
                    "actual_qty": e.actual_qty,
                    "excess": round(e.actual_qty - b.total_estimated_qty, 4),
                    "justification": e.justification,
                })
    result.sort(key=lambda x: (x["project"], x["month"], x["activity"], x["floor"]))
    return result


# ─── Compare ─────────────────────────────────────────────────────────────────

def get_comparison(db: Session, activity_id: int, month: str) -> dict:
    """Cross-project comparison for a single activity and month."""
    act = db.query(QsActivity).filter_by(id=activity_id).first()
    if not act:
        raise HTTPException(404, "Activity not found")

    projects = db.query(QsProject).order_by(QsProject.name).all()
    rows = []
    for proj in projects:
        floors = (
            db.query(QsFloor)
            .filter_by(project_id=proj.id)
            .order_by(QsFloor.display_order)
            .all()
        )
        cells = []
        for floor in floors:
            baseline = db.query(QsBaseline).filter_by(
                project_id=proj.id, activity_id=activity_id, floor_id=floor.id
            ).first()
            entry = db.query(QsMonthlyEntry).filter_by(
                project_id=proj.id, activity_id=activity_id,
                floor_id=floor.id, month=month,
            ).first()
            cells.append({
                "floor": floor.name,
                "total_estimated_qty": baseline.total_estimated_qty if baseline else None,
                "estimate_actual_qty_till_date": entry.estimate_actual_qty_till_date if entry else None,
                "actual_qty": entry.actual_qty if entry else None,
                "overrun": (
                    entry.actual_qty > baseline.total_estimated_qty
                    if entry and baseline else False
                ),
            })
        rows.append({"project": proj.name, "cells": cells})

    return {"activity": _activity_to_dict(act), "month": month, "projects": rows}


# ─── Baseline history ─────────────────────────────────────────────────────────

def get_baseline_history(db: Session, project_id: int, activity_id: int, floor_id: int) -> list[dict]:
    baseline = db.query(QsBaseline).filter_by(
        project_id=project_id, activity_id=activity_id, floor_id=floor_id
    ).first()
    if not baseline:
        return []
    logs = (
        db.query(QsBaselineChangeLog)
        .filter_by(baseline_id=baseline.id)
        .order_by(QsBaselineChangeLog.changed_at)
        .all()
    )
    return [
        {
            "old_value": l.old_value,
            "new_value": l.new_value,
            "reason": l.reason,
            "changed_by": l.changed_by,
            "changed_at": l.changed_at.isoformat(),
        }
        for l in logs
    ]


# ─── Excel Export ─────────────────────────────────────────────────────────────

_BLUE = "FF0066CC"
_LIGHT_BLUE = "FFD6E4F7"
_RED_FILL = "FFFFC7CE"
_RED_FONT = "FF9C0006"
_GREY = "FFF2F2F2"
_BORDER_COLOR = "FFD0D0D0"

_thin = Side(style="thin", color=_BORDER_COLOR)
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _header_font(color: str = "FFFFFFFF", bold: bool = True) -> Font:
    return Font(bold=bold, color=color, size=10)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def export_to_excel(db: Session, project_id: int, month: str) -> bytes:
    data = get_project_data(db, project_id, month)
    activities: list[dict] = data["activities"]
    floors: list[dict] = data["floors"]
    cells_map: dict[tuple, dict] = {
        (c["activity_id"], c["floor_id"]): c for c in data["cells"]
    }

    wb = Workbook()
    ws = wb.active
    ws.title = f"QS {month}"

    # ── Header row 1: project + month banner ────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=2 + len(floors) * 3 + 1)
    banner = ws.cell(1, 1)
    banner.value = f"{data['project']['name']} — Quantity Sheet — {month}"
    banner.font = _header_font()
    banner.fill = _fill(_BLUE)
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Header row 2: Activity | Unit | Floor… (3 cols each) | Justification ─
    ws.cell(2, 1, "Activity").font = _header_font(color="FF000000")
    ws.cell(2, 1).fill = _fill(_GREY)
    ws.cell(2, 1).border = _border
    ws.cell(2, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.cell(2, 2, "Unit").font = _header_font(color="FF000000")
    ws.cell(2, 2).fill = _fill(_GREY)
    ws.cell(2, 2).border = _border
    ws.cell(2, 2).alignment = Alignment(horizontal="center", vertical="center")

    col = 3
    for floor in floors:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 2)
        cell = ws.cell(2, col)
        cell.value = floor["name"]
        cell.font = _header_font()
        cell.fill = _fill(_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border
        col += 3

    justification_col = col
    ws.cell(2, justification_col, "Justification").font = _header_font(color="FF000000")
    ws.cell(2, justification_col).fill = _fill(_GREY)
    ws.cell(2, justification_col).border = _border
    ws.cell(2, justification_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 20

    # ── Header row 3: sub-headers per floor ──────────────────────────────────
    ws.cell(3, 1).border = _border
    ws.cell(3, 2).border = _border
    ws.cell(3, justification_col).border = _border

    col = 3
    sub_labels = ["Total Est. Qty\n(As per Drawing)", "Est. Actual Qty\nTill Date", "Actual Qty"]
    for _ in floors:
        for label in sub_labels:
            c = ws.cell(3, col)
            c.value = label
            c.font = Font(bold=True, color="FF000000", size=8)
            c.fill = _fill(_LIGHT_BLUE)
            c.border = _border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            col += 1
    ws.row_dimensions[3].height = 30

    # ── Data rows ────────────────────────────────────────────────────────────
    for row_idx, act in enumerate(activities, start=4):
        ws.cell(row_idx, 1, act["name"]).border = _border
        ws.cell(row_idx, 1).alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row_idx, 2, act["unit"]).border = _border
        ws.cell(row_idx, 2).alignment = Alignment(horizontal="center", vertical="center")

        justifications = []
        col = 3
        for floor in floors:
            c = cells_map.get((act["id"], floor["id"]), {})
            total_est = c.get("total_estimated_qty")
            est_actual = c.get("estimate_actual_qty_till_date")
            actual = c.get("actual_qty")
            justification = c.get("justification") or ""

            vals = [total_est, est_actual, actual]
            is_overrun = (
                actual is not None and total_est is not None and actual > total_est
            )
            for v in vals:
                cell = ws.cell(row_idx, col)
                cell.value = v
                cell.border = _border
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if is_overrun:
                    cell.font = Font(color=_RED_FONT, size=10)
                    cell.fill = _fill(_RED_FILL)
                col += 1

            if is_overrun and justification:
                justifications.append(justification)

        just_cell = ws.cell(row_idx, justification_col)
        just_cell.value = "; ".join(justifications) if justifications else None
        just_cell.border = _border
        just_cell.alignment = Alignment(vertical="center", wrap_text=True)

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(1)].width = 28  # Activity
    ws.column_dimensions[get_column_letter(2)].width = 8   # Unit
    for i in range(len(floors) * 3):
        ws.column_dimensions[get_column_letter(3 + i)].width = 14
    ws.column_dimensions[get_column_letter(justification_col)].width = 36

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
