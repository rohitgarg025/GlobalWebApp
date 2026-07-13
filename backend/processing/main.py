import pandas as pd
import os
import time
from datetime import datetime

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

START_ROW = 3

LABOUR_REPORT_TITLE = "Labour Costing Report"
LABOUR_SUMMARY_TITLE = 'Labour Costing Summary'

# Report type constants
LABOUR_REPORT = "Labour Report"
MATERIAL_REPORT = "Material Reconciliation Report"
ACTIVITY_COSTING_REPORT = "Activity Wise Costing Report"
ALL_REPORTS = "All of the above"
MULTIPLE_COST_REPORTS = "Multiple Projects - Cost Reports"
MONTHWISE_SITEWISE_LABOUR_QTY_REPORT = "Monthwise Sitewise Labour Quantity Report"
THEORETICAL_CONSUMPTION_REPORT = "Theoretical Consumption Report"
FUND_REPORT = "Fund Report"
BILL_MASTER_REPORT = "Bill Master Report"
MONTHLY_PROJECTWISE_LABOUR_REPORT = "Monthly Projectwise Labour Report"
COMPANY_TITLE = "Global Buildestate Projects Pvt. Ltd."

def format_excel_with_headers(writer, sheet_name, df, report_title, project_name=None):
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    worksheet['A1'] = f'Report Type : {report_title}'
    worksheet['A1'].font = Font(name='Arial', size=14, bold=True)
    worksheet['A1'].alignment = Alignment(horizontal='left')

    if project_name:
        worksheet['A2'] = f'Project Name : {project_name}'
        worksheet['A2'].font = Font(name='Arial', size=11, bold=True)
        worksheet['A2'].alignment = Alignment(horizontal='left')

    row_offset = 3 if project_name else 2
    worksheet[f'A{row_offset}'] = f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    worksheet[f'A{row_offset}'].font = Font(name='Arial', size=9, italic=True)

    header_row = row_offset + 1
    header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def get_output_excel(project_name, base_filename, file_extension, output_dir=""):
    project_name = project_name.lower().replace(' ', '_') if project_name else 'project'
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    new_filename = f"{project_name}_{base_filename}_{timestamp}{file_extension}"
    if output_dir:
        return os.path.join(output_dir, new_filename)
    return new_filename


def is_wbs_contain_string(df, str):
    return df['Parent WBS Name'].str.contains(str, case=False, na=False)


def get_labour_amt_summary_list(df, amt_type):
    flag_general_site_work = is_wbs_contain_string(df, "general site work")
    flag_site_expenses = is_wbs_contain_string(df, "site expenses") | is_wbs_contain_string(df, "site administrative expenses")
    flag_hire_rent = is_wbs_contain_string(df, "hire & rent")
    flag_site_infra_work = is_wbs_contain_string(df, "site infra work")

    general_site_work = float(df.loc[flag_general_site_work, amt_type].sum() or 0)
    site_expenses = float(df.loc[flag_site_expenses, amt_type].sum() or 0)
    hire_rent = float(df.loc[flag_hire_rent, amt_type].sum() or 0)
    site_infra_work = float(df.loc[flag_site_infra_work, amt_type].sum() or 0)

    overhead_total = site_expenses + hire_rent
    productive_total = float(df[amt_type].sum()) - (general_site_work + overhead_total + site_infra_work)

    return [productive_total, site_infra_work, general_site_work, hire_rent, site_expenses]


def get_consolidated_report(df_list):
    file1 = df_list[0].copy()
    file2 = df_list[1].copy()
    file3 = pd.DataFrame()
    if len(df_list) > 2:
        file3 = df_list[2].copy()

    file1['Key'] = file1['Activity Name'].astype(str).str.strip() + "   " + file1['Item Desc'].astype(str).str.strip()
    file2['Key'] = file2['Activity Name'].astype(str).str.strip() + "   " + file2['Resource Name'].astype(str).str.strip()

    file1 = file1.merge(file2[['Key', 'Resource Type']], on='Key', how='left')
    return [file1, file2, file3]


def clean_df(df):
    try:
        df = df.rename(columns={
            "3Est Qty": "Est Qty",
            "3Theoritical Qty": "Theoritical Qty",
            "3Actual Qty": "Actual Qty",
            "3Difference Qty": "Difference Qty",
        })
        return df
    except Exception as e:
        print(f"Invalid columns: {e}")
        return None


def get_merged_df_with_stock_report(pivot, df_stock_report, stock_cols):
    pivot = pivot.merge(df_stock_report[['Item Desc'] + stock_cols], on='Item Desc', how='left')
    return pivot


def get_labour_sheet(df_list, output_dir=""):
    file_list = get_consolidated_report(df_list)
    df = file_list[0].copy()
    df = df[df['Resource Type'].str.lower() == 'service']
    df = clean_df(df)

    df = df.sort_values(by='Parent WBS Name')

    estimated_amt_labour_summary_list = get_labour_amt_summary_list(df, "Est Amt")
    theoritical_amt_labour_summary_list = get_labour_amt_summary_list(df, "Theoritical Amt")
    actual_amt_labour_summary_list = get_labour_amt_summary_list(df, "Actual Amt")

    summary = pd.DataFrame({
        'Labour Costing Summary': [
            'Productive Labour',
            'Site Infra Work',
            'General Site Work',
            'Hire & Rent',
            'Site Admin Expenses'
        ],
        'Estimated Amount': estimated_amt_labour_summary_list,
        'Theoritical Amount': theoritical_amt_labour_summary_list,
        'Actual Amount': actual_amt_labour_summary_list
    })

    project_name = df['Project Name'].iloc[0] if 'Project Name' in df.columns else None
    columns_to_delete = ["Sr", "Activity Code", "Project Name", "Sub Project", "Activity Unit", "3Total Qty", "Key", "Resource Type"]
    df = df.drop(columns=[c for c in columns_to_delete if c in df.columns])

    df = df.rename(columns={
        'Theoritical Qty': 'DPR Work Qty',
        'Actual Qty': 'SRN Billed Qty',
        'Theoritical Amt': 'Theoritical Work Amt',
        'Actual Amt': 'SRN Amt'
    })

    df["Act Rate"] = (df["SRN Amt"] / df["SRN Billed Qty"]).where(df["SRN Billed Qty"] != 0, 0)
    df["Balance Work Amt [Est - Th]"] = df["Est Amt"] - df["Theoritical Work Amt"]
    df["Percentage Work Completed [DPR/Est]*100"] = ((df["DPR Work Qty"] / df["Est Qty"]).where(df["Est Qty"] != 0, 0)) * 100
    df["Work Amt"] = df["DPR Work Qty"] * df["Act Rate"]

    output_path_file = get_output_excel(project_name, "Labour_Report", ".xlsx", output_dir)
    with pd.ExcelWriter(output_path_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Detailed Data', startrow=START_ROW)
        summary.to_excel(writer, index=False, sheet_name='Summary', startrow=START_ROW)
        format_excel_with_headers(writer, 'Detailed Data', df, LABOUR_REPORT_TITLE, project_name)
        format_excel_with_headers(writer, 'Summary', summary, LABOUR_SUMMARY_TITLE, project_name)

    print(f"Labour report generated: {output_path_file}")
    return output_path_file


def get_material_sheet(df_list, output_dir=""):
    file_list = get_consolidated_report(df_list)
    df = file_list[0].copy()
    df = df[df['Resource Type'].str.lower() == 'material']
    df = clean_df(df)
    project_name = df['Project Name'].iloc[0] if 'Project Name' in df.columns else None

    numeric_cols = ["Est Qty", "Est Rate", "Est Amt", "Theoritical Qty", "Actual Qty", "Theoritical Amt", "Actual Amt"]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    pivot = pd.pivot_table(
        df,
        index=["Item Group", "Item Desc", "Resource Unit"],
        values={"Est Qty": "sum", "Est Rate": "mean", "Est Amt": "sum",
                "Theoritical Qty": "sum", "Theoritical Amt": "sum",
                "Actual Qty": "sum", "Actual Amt": "sum"},
        aggfunc={"Est Qty": "sum", "Est Rate": "mean", "Est Amt": "sum",
                 "Theoritical Qty": "sum", "Theoritical Amt": "sum",
                 "Actual Qty": "sum", "Actual Amt": "sum"},
        fill_value=0,
        margins=True,
        margins_name="Total"
    ).reset_index()

    pivot["Wastage"] = pivot["Actual Qty"] - pivot["Theoritical Qty"]
    pivot["Wastage %"] = ((pivot["Wastage"] / pivot["Theoritical Qty"]) * 100).where(pivot["Theoritical Qty"] != 0, 0)
    pivot["Actual Rate"] = (pivot["Actual Amt"] / pivot["Actual Qty"]).where(pivot["Actual Qty"] != 0, 0)

    desired_columns_order = [
        "Item Group", "Item Desc", "Resource Unit",
        "Est Qty", "Est Rate", "Est Amt",
        "Theoritical Qty", "Theoritical Amt",
        "Actual Qty", "Actual Rate", "Actual Amt",
        "Wastage", "Wastage %"
    ]

    stock_cols = []
    if len(file_list) > 2 and not file_list[2].empty:
        df_stock_report = file_list[2].copy()
        stock_cols = ['Stock Received Qty (B)', 'Project Transfer Qty (E)', 'Stock Issue Qty (D)', 'Closing Stock Qty (CG)']
        pivot = get_merged_df_with_stock_report(pivot.copy(), df_stock_report, stock_cols)
        for col in stock_cols:
            pivot[col] = pd.to_numeric(pivot[col], errors='coerce').fillna(0)
        desired_columns_order.extend(stock_cols)

    pivot = pivot.reindex(columns=desired_columns_order)

    for col in ["Est Qty", "Est Rate", "Est Amt", "Theoritical Qty", "Theoritical Amt",
                "Actual Qty", "Actual Rate", "Actual Amt"]:
        pivot[col] = pivot[col].round(2)

    if stock_cols:
        for col in stock_cols:
            pivot[col] = pivot[col].round(2)
        pivot["Actual - Tag Issue Qty Difference"] = pivot["Stock Issue Qty (D)"] - pivot["Actual Qty"]

    pivot = pivot.sort_values(by=["Item Group", "Item Desc"])

    output_path = get_output_excel(project_name, "Material_Reconciliation_Report", ".xlsx", output_dir)
    pivot.to_excel(output_path, index=False)
    print(f"Material Reconciliation Report generated: {output_path}")
    return output_path


def get_activity_costing_report(df_list, output_dir=""):
    file_list = get_consolidated_report(df_list)
    df = file_list[0].copy()
    df = clean_df(df)
    project_name = df['Project Name'].iloc[0] if 'Project Name' in df.columns else None

    numeric_cols = ["Est Amt", "Theoritical Amt", "Actual Amt"]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    pivot = pd.pivot_table(
        df,
        index=["Parent WBS Name", "Activity Name", "Activity Unit"],
        columns=["Resource Type"],
        values=numeric_cols,
        aggfunc=sum,
        fill_value=0,
        margins=True,
        margins_name="Total"
    )

    new_columns = [f"{level1.upper()} - {level2.upper()}" for level1, level2 in pivot.columns]
    pivot.columns = new_columns

    key_cols = ["Parent WBS Name", "Activity Name", "Activity Unit"]
    pivot = pivot.reset_index()

    amount_order = ["Est Amt", "Theoritical Amt", "Actual Amt"]
    resource_order = ["MATERIAL", "SERVICE", "TOTAL"]

    desired_data_cols = []
    for amount in amount_order:
        for resource in resource_order:
            col_name = f"{amount.upper()} - {resource.upper()}"
            if col_name in pivot.columns:
                desired_data_cols.append(col_name)

    final_col_order = key_cols + desired_data_cols
    pivot = pivot.reindex(columns=final_col_order)

    numeric_cols_to_round = [col for col in pivot.columns if col not in key_cols]
    for col in numeric_cols_to_round:
        pivot[col] = pd.to_numeric(pivot[col], errors='coerce').fillna(0)
        pivot[col] = pivot[col].round(2)

    output_path = get_output_excel(project_name, "Activity_Costing_Report", ".xlsx", output_dir)
    pivot.to_excel(output_path, index=False)
    print(f"Activity Costing Report generated: {output_path}")
    return output_path


def get_all_reports_single_project(df_list, output_dir=""):
    ls = get_labour_sheet(df_list.copy(), output_dir)
    ms = get_material_sheet(df_list.copy(), output_dir)
    acs = get_activity_costing_report(df_list.copy(), output_dir)
    return [ls, ms, acs]


def get_multi_project_list(df_list):
    return [df_list[i:i + 3] for i in range(0, len(df_list), 3)]


def excel_transform(df_list, transform_option, output_dir=""):
    """
    Main dispatcher. Returns a list of output file paths.
    """
    if transform_option == LABOUR_REPORT:
        return [get_labour_sheet(df_list, output_dir)]
    elif transform_option == MATERIAL_REPORT:
        return [get_material_sheet(df_list, output_dir)]
    elif transform_option == ACTIVITY_COSTING_REPORT:
        return [get_activity_costing_report(df_list, output_dir)]
    elif transform_option == ALL_REPORTS:
        return get_all_reports_single_project(df_list, output_dir)
    elif transform_option == THEORETICAL_CONSUMPTION_REPORT:
        return [theoretical_consumption_sheet(df_list, output_dir)]
    elif transform_option == FUND_REPORT:
        return [fund_report_sheet(df_list, output_dir)]
    elif transform_option == BILL_MASTER_REPORT:
        return [bill_master_report_sheet(df_list, output_dir)]
    elif transform_option == MONTHLY_PROJECTWISE_LABOUR_REPORT:
        return [monthly_projectwise_labour_sheet(df_list, output_dir)]
    elif transform_option == MULTIPLE_COST_REPORTS:
        multi_project_sheet_list = get_multi_project_list(df_list)
        all_outputs = []
        for project_files in multi_project_sheet_list:
            all_outputs.extend(get_all_reports_single_project(project_files, output_dir))
            time.sleep(1)
        return all_outputs
    else:
        raise ValueError(f"Invalid transform option: {transform_option}")


def theoretical_consumption_sheet(df_list, output_dir=""):
    file_list = get_consolidated_report(df_list)
    df = file_list[0].copy()
    df = df[df['Resource Type'].str.lower() == 'material']
    df = clean_df(df)
    # Some ERP exports prefix the activity total with "3" like the other qty columns
    df = df.rename(columns={"3Total Qty": "Total Qty"})
    project_name = df['Project Name'].iloc[0] if 'Project Name' in df.columns else None

    numeric_cols = ["Total Qty", "Est Qty", "Theoritical Qty"]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    # Resource Coefficient = material qty estimated per unit of activity
    df["Resource Coefficient"] = (df["Est Qty"] / df["Total Qty"]).where(df["Total Qty"] != 0, 0)

    desired_columns_order = [
        "Activity Name", "Activity Unit", "Total Qty",
        "Item Group", "Item Desc", "Resource Coefficient",
        "Resource Unit", "Est Qty", "Theoritical Qty",
    ]
    df = df.reindex(columns=desired_columns_order)

    df = df.sort_values(by=["Activity Name", "Item Group"]).reset_index(drop=True)

    for col in ["Total Qty", "Est Qty", "Theoritical Qty"]:
        df[col] = df[col].round(2)
    df["Resource Coefficient"] = df["Resource Coefficient"].round(6)

    sheet_name = 'Theoretical Consumption'
    output_path_file = get_output_excel(project_name, "Theoretical_Consumption_Report", ".xlsx", output_dir)

    with pd.ExcelWriter(output_path_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=START_ROW)
        format_excel_with_headers(writer, sheet_name, df, THEORETICAL_CONSUMPTION_REPORT, project_name)

        worksheet = writer.sheets[sheet_name]

        # Detail data occupies rows first_data_row .. last_data_row (1-indexed);
        # to_excel(startrow=START_ROW) puts the header on excel row START_ROW+1.
        first_data_row = START_ROW + 2
        last_data_row = first_data_row + len(df) - 1
        item_desc_col = get_column_letter(df.columns.get_loc("Item Desc") + 1)
        th_qty_col = get_column_letter(df.columns.get_loc("Theoritical Qty") + 1)
        criteria_range = f"${item_desc_col}${first_data_row}:${item_desc_col}${last_data_row}"
        sum_range = f"${th_qty_col}${first_data_row}:${th_qty_col}${last_data_row}"

        # ── Material-wise summary (two blank rows below the detail table) ──────
        summary_start = last_data_row + 3

        title_cell = worksheet.cell(row=summary_start, column=1)
        title_cell.value = 'Material Wise Theoretical Consumption Summary'
        title_cell.font = Font(name='Arial', size=12, bold=True)

        header_row = summary_start + 1
        summary_headers = ['Item Desc', 'Resource Unit', 'Total Theoritical Qty']
        header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for col_idx, header in enumerate(summary_headers, start=1):
            cell = worksheet.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # One summary row per material with a live =SUMIF formula; the criteria
        # references the Item Desc cell of the summary row itself, so clicking
        # the total in Excel highlights the detail ranges it draws from.
        write_row = header_row + 1
        for item_desc, group in df.groupby("Item Desc", sort=True):
            worksheet.cell(row=write_row, column=1).value = item_desc
            worksheet.cell(row=write_row, column=2).value = group["Resource Unit"].iloc[0]
            worksheet.cell(row=write_row, column=3).value = (
                f"=SUMIF({criteria_range},A{write_row},{sum_range})"
            )
            write_row += 1

    print(f"Theoretical Consumption Report generated: {output_path_file}")
    return output_path_file


def _write_fund_section(worksheet, df, start_row, title, columns, header_fill, header_font, header_alignment):
    """Write a titled section (title row + header row + data rows) into worksheet."""
    # Section title
    title_cell = worksheet.cell(row=start_row, column=1)
    title_cell.value = title
    title_cell.font = Font(name='Arial', size=12, bold=True)

    # Column headers
    header_row = start_row + 1
    for col_idx, col_name in enumerate(columns, start=1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Data rows
    data_start = header_row + 1
    for row_idx, (_, row) in enumerate(df.iterrows()):
        for col_idx, col_name in enumerate(columns, start=1):
            worksheet.cell(row=data_start + row_idx, column=col_idx).value = row.get(col_name)

    return data_start + len(df)  # row index after the last data row


def _write_summary_section(worksheet, df, start_row, title, header_fill, header_font, header_alignment):
    """Write a summary section (title + header + grouped data) into worksheet."""
    summary_cols = ['Item Group', 'Item Desc', 'Resource Unit', 'Total Est Qty', 'Total Est Amt']

    title_cell = worksheet.cell(row=start_row, column=1)
    title_cell.value = title
    title_cell.font = Font(name='Arial', size=12, bold=True)

    header_row = start_row + 1
    for col_idx, col_name in enumerate(summary_cols, start=1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    data_start = header_row + 1
    for row_idx, (_, row) in enumerate(df.iterrows()):
        worksheet.cell(row=data_start + row_idx, column=1).value = row.get('Item Group')
        worksheet.cell(row=data_start + row_idx, column=2).value = row.get('Item Desc')
        worksheet.cell(row=data_start + row_idx, column=3).value = row.get('Resource Unit')
        worksheet.cell(row=data_start + row_idx, column=4).value = row.get('Total Est Qty')
        worksheet.cell(row=data_start + row_idx, column=5).value = row.get('Total Est Amt')

    return data_start + len(df)


def fund_report_sheet(df_list, output_dir=""):
    file_list = get_consolidated_report(df_list)
    df = file_list[0].copy()
    df = clean_df(df)
    df = df.rename(columns={"3Total Qty": "Total Qty"})
    project_name = df['Project Name'].iloc[0] if 'Project Name' in df.columns else None

    numeric_cols = ["Total Qty", "Est Qty", "Est Rate", "Est Amt"]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df["Resource Coefficient"] = (df["Est Qty"] / df["Total Qty"]).where(df["Total Qty"] != 0, 0)

    desired_columns = [
        "Parent WBS Name", "Activity Name", "Activity Unit", "Total Qty",
        "Resource Type", "Item Group", "Item Desc", "Resource Unit",
        "Resource Coefficient", "Est Qty", "Est Rate", "Est Amt",
    ]
    df = df.reindex(columns=desired_columns)

    df = df.sort_values(
        by=["Parent WBS Name", "Activity Name", "Resource Type", "Item Group", "Item Desc"]
    ).reset_index(drop=True)

    for col in ["Total Qty", "Est Qty", "Est Rate", "Est Amt"]:
        df[col] = df[col].round(2)
    df["Resource Coefficient"] = df["Resource Coefficient"].round(6)

    item_group = df['Item Group'].fillna('').str.lower()
    parent_wbs = df['Parent WBS Name'].fillna('').str.strip().str.lower()
    activity_name = df['Activity Name'].fillna('').str.strip().str.lower()
    mask_site_infra = (parent_wbs == 'site infra work') | activity_name.str.startswith('infra')
    mask_non_productive = ~mask_site_infra & item_group.str.contains('labour-general works', na=False)
    mask_overheads = ~mask_site_infra & ~mask_non_productive & (
        item_group.str.contains('site expenses', na=False) | item_group.str.contains('hire & rent', na=False)
    )
    mask_productive = ~mask_site_infra & ~mask_non_productive & ~mask_overheads

    df_productive = df[mask_productive].reset_index(drop=True)
    df_site_infra = df[mask_site_infra].reset_index(drop=True)
    df_non_productive = df[mask_non_productive].reset_index(drop=True)
    df_overheads = df[mask_overheads].reset_index(drop=True)

    header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    output_path_file = get_output_excel(project_name, "Fund_Report", ".xlsx", output_dir)

    with pd.ExcelWriter(output_path_file, engine='openpyxl') as writer:
        # ── Sheet 1: Fund Report ──────────────────────────────────────────────
        sheet1_name = 'Fund Report'
        # Bootstrap the sheet by writing an empty DataFrame
        pd.DataFrame(columns=desired_columns).to_excel(
            writer, index=False, sheet_name=sheet1_name, startrow=0
        )
        ws1 = writer.sheets[sheet1_name]

        # Report header (rows 1-3)
        ws1['A1'] = f'Report Type : {FUND_REPORT}'
        ws1['A1'].font = Font(name='Arial', size=14, bold=True)
        ws1['A2'] = f'Project Name : {project_name}' if project_name else ''
        ws1['A2'].font = Font(name='Arial', size=11, bold=True)
        ws1['A3'] = f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws1['A3'].font = Font(name='Arial', size=9, italic=True)

        current_row = START_ROW + 1  # row 4 — first section starts here

        current_row = _write_fund_section(
            ws1, df_productive, current_row,
            "Productive Work", desired_columns,
            header_fill, header_font, header_alignment
        )
        current_row += 2  # two blank rows

        current_row = _write_fund_section(
            ws1, df_site_infra, current_row,
            "Site Infra Work", desired_columns,
            header_fill, header_font, header_alignment
        )
        current_row += 2

        current_row = _write_fund_section(
            ws1, df_non_productive, current_row,
            "Non-Productive (General Site Work)", desired_columns,
            header_fill, header_font, header_alignment
        )
        current_row += 2

        _write_fund_section(
            ws1, df_overheads, current_row,
            "Overheads", desired_columns,
            header_fill, header_font, header_alignment
        )

        # ── Sheet 2: Summary ─────────────────────────────────────────────────
        sheet2_name = 'Summary'
        pd.DataFrame().to_excel(writer, index=False, sheet_name=sheet2_name, startrow=0)
        ws2 = writer.sheets[sheet2_name]

        ws2['A1'] = f'Report Type : {FUND_REPORT} — Summary'
        ws2['A1'].font = Font(name='Arial', size=14, bold=True)
        ws2['A2'] = f'Project Name : {project_name}' if project_name else ''
        ws2['A2'].font = Font(name='Arial', size=11, bold=True)
        ws2['A3'] = f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws2['A3'].font = Font(name='Arial', size=9, italic=True)

        def make_summary(source_df, resource_type):
            subset = source_df[
                source_df['Resource Type'].str.lower() == resource_type
            ].copy()
            if subset.empty:
                return pd.DataFrame(columns=['Item Group', 'Item Desc', 'Resource Unit', 'Total Est Qty', 'Total Est Amt'])
            grouped = (
                subset.groupby(['Item Group', 'Item Desc', 'Resource Unit'], as_index=False)
                .agg({"Est Qty": "sum", "Est Amt": "sum"})
                .rename(columns={"Est Qty": "Total Est Qty", "Est Amt": "Total Est Amt"})
            )
            grouped['Total Est Qty'] = grouped['Total Est Qty'].round(2)
            grouped['Total Est Amt'] = grouped['Total Est Amt'].round(2)
            return grouped

        mat_summary = make_summary(df_productive, 'material')
        svc_summary = make_summary(df_productive, 'service')

        summary_row = START_ROW + 1  # row 4

        summary_row = _write_summary_section(
            ws2, mat_summary, summary_row,
            "Material Wise Summary (Productive Work)",
            header_fill, header_font, header_alignment
        )
        summary_row += 2

        _write_summary_section(
            ws2, svc_summary, summary_row,
            "Service Wise Summary (Productive Work)",
            header_fill, header_font, header_alignment
        )

    print(f"Fund Report generated: {output_path_file}")
    return output_path_file


def _parse_srn_export(raw):
    """Locate the header row of a raw SRN Report export (variable-length preamble)
    and return a clean data DataFrame with numeric qty/amount columns."""
    header_idx = None
    for idx, value in raw.iloc[:, 0].items():
        if str(value).strip().lower() == "sr":
            header_idx = idx
            break
    if header_idx is None:
        raise ValueError("Could not find the SRN Report header row (first column 'Sr').")

    df = raw.iloc[header_idx + 1:].copy()
    df.columns = [str(c).strip() for c in raw.iloc[header_idx]]
    df = df.reset_index(drop=True)

    df = df[df['Particulars'].notna()]
    df['Group'] = df['Group'].fillna('')
    df['Unit'] = df['Unit'].fillna('')
    for col in ['SRN Qty', 'SRN Amount']:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    return df


def bill_master_report_sheet(df_list, output_dir=""):
    """Contractor-wise billing matrix built from a raw SRN Report export."""
    df = _parse_srn_export(df_list[0].copy())
    df = df[df['Contractor Name'].notna()]

    project_name = None
    if 'Project Name' in df.columns and df['Project Name'].notna().any():
        project_name = df['Project Name'].dropna().iloc[0]

    pivot = df.pivot_table(
        index=['Group', 'Particulars', 'Unit'],
        columns='Contractor Name',
        values=['SRN Qty', 'SRN Amount'],
        aggfunc='sum',
    )

    contractors = sorted(df['Contractor Name'].unique())
    index_headers = ['Item Group', 'Particulars', 'Unit']
    sub_headers = ['Billed Qty', 'Billed Rate', 'Billed Amount']

    pivot = pivot.sort_index()
    row_keys = list(pivot.index)

    output_path_file = get_output_excel(project_name, "Bill_Master_Report", ".xlsx", output_dir)

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Bill Master'

    # ── Styles ───────────────────────────────────────────────────────────────
    band_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    band_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    sub_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    sub_font = Font(name='Arial', size=9, bold=True, color='1F3864')
    total_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    total_font = Font(name='Arial', size=10, bold=True)
    alt_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
    data_font = Font(name='Arial', size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')
    thin = Side(style='thin', color='B7C4D6')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    NUM_FMT = '#,##0.00'

    # ── Report header ────────────────────────────────────────────────────────
    ws['A1'] = f'Report Type : {BILL_MASTER_REPORT}'
    ws['A1'].font = Font(name='Arial', size=14, bold=True)
    if project_name:
        ws['A2'] = f'Project Name : {project_name}'
        ws['A2'].font = Font(name='Arial', size=11, bold=True)
    ws['A3'] = f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A3'].font = Font(name='Arial', size=9, italic=True)

    band_row = 5
    sub_row = 6
    first_data_row = 7

    # ── Header band: index headers merged vertically ─────────────────────────
    for col_idx, header in enumerate(index_headers, start=1):
        ws.merge_cells(start_row=band_row, start_column=col_idx, end_row=sub_row, end_column=col_idx)
        cell = ws.cell(row=band_row, column=col_idx)
        cell.value = header
        cell.fill = band_fill
        cell.font = band_font
        cell.alignment = center
        for r in (band_row, sub_row):
            ws.cell(row=r, column=col_idx).border = border

    # ── Header band: contractor blocks (3 columns each) + Grand Total ────────
    blocks = contractors + ['Grand Total']
    for block_no, contractor in enumerate(blocks):
        start_col = len(index_headers) + 1 + block_no * 3
        ws.merge_cells(start_row=band_row, start_column=start_col, end_row=band_row, end_column=start_col + 2)
        cell = ws.cell(row=band_row, column=start_col)
        cell.value = contractor
        cell.fill = band_fill
        cell.font = band_font
        cell.alignment = center
        for offset, sub in enumerate(sub_headers):
            sub_cell = ws.cell(row=sub_row, column=start_col + offset)
            sub_cell.value = sub
            sub_cell.fill = sub_fill
            sub_cell.font = sub_font
            sub_cell.alignment = center
            sub_cell.border = border
            ws.cell(row=band_row, column=start_col + offset).border = border

    # ── Data rows ─────────────────────────────────────────────────────────────
    def write_cell(row, col, value, *, fmt=None, font=data_font, fill=None, align=right):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.font = font
        cell.alignment = align
        cell.border = border
        if fmt:
            cell.number_format = fmt
        if fill:
            cell.fill = fill
        return cell

    grand_total_qty = {c: 0.0 for c in contractors}
    grand_total_amt = {c: 0.0 for c in contractors}
    overall_qty_total = 0.0
    overall_amt_total = 0.0

    for row_no, (group, particulars, unit) in enumerate(row_keys):
        excel_row = first_data_row + row_no
        row_fill = alt_fill if row_no % 2 == 1 else None

        write_cell(excel_row, 1, group, font=data_font, fill=row_fill, align=left)
        write_cell(excel_row, 2, particulars, font=data_font, fill=row_fill, align=left)
        write_cell(excel_row, 3, unit, font=data_font, fill=row_fill, align=center)

        row_qty_total = 0.0
        row_amt_total = 0.0
        for block_no, contractor in enumerate(contractors):
            start_col = len(index_headers) + 1 + block_no * 3
            try:
                qty = pivot.loc[(group, particulars, unit), ('SRN Qty', contractor)]
            except KeyError:
                qty = None
            try:
                amt = pivot.loc[(group, particulars, unit), ('SRN Amount', contractor)]
            except KeyError:
                amt = None
            qty = None if pd.isna(qty) else float(qty)
            amt = None if pd.isna(amt) else float(amt)

            if qty is None and amt is None:
                for offset in range(3):
                    write_cell(excel_row, start_col + offset, None, fill=row_fill)
                continue

            qty = qty or 0.0
            amt = amt or 0.0
            rate = (amt / qty) if qty else 0.0
            write_cell(excel_row, start_col, round(qty, 2), fmt=NUM_FMT, fill=row_fill)
            write_cell(excel_row, start_col + 1, round(rate, 2), fmt=NUM_FMT, fill=row_fill)
            write_cell(excel_row, start_col + 2, round(amt, 2), fmt=NUM_FMT, fill=row_fill)

            row_qty_total += qty
            row_amt_total += amt
            grand_total_qty[contractor] += qty
            grand_total_amt[contractor] += amt

        # Grand Total block for this row
        gt_col = len(index_headers) + 1 + len(contractors) * 3
        gt_rate = (row_amt_total / row_qty_total) if row_qty_total else 0.0
        write_cell(excel_row, gt_col, round(row_qty_total, 2), fmt=NUM_FMT, font=total_font, fill=row_fill)
        write_cell(excel_row, gt_col + 1, round(gt_rate, 2), fmt=NUM_FMT, font=total_font, fill=row_fill)
        write_cell(excel_row, gt_col + 2, round(row_amt_total, 2), fmt=NUM_FMT, font=total_font, fill=row_fill)
        overall_qty_total += row_qty_total
        overall_amt_total += row_amt_total

    # ── Bottom total row (amounts only — quantities mix units) ───────────────
    total_row = first_data_row + len(row_keys)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    cell = ws.cell(row=total_row, column=1)
    cell.value = 'Total Billed Amount'
    cell.font = total_font
    cell.fill = total_fill
    cell.alignment = left
    for c in range(1, 4):
        ws.cell(row=total_row, column=c).border = border
        ws.cell(row=total_row, column=c).fill = total_fill
    for block_no, contractor in enumerate(blocks):
        start_col = len(index_headers) + 1 + block_no * 3
        if contractor == 'Grand Total':
            amt = overall_amt_total
        else:
            amt = grand_total_amt[contractor]
        for offset in range(3):
            write_cell(total_row, start_col + offset, None, font=total_font, fill=total_fill)
        write_cell(total_row, start_col + 2, round(amt, 2), fmt=NUM_FMT, font=total_font, fill=total_fill)

    # ── Layout polish ─────────────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=first_data_row, column=len(index_headers) + 1)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 10
    for block_no in range(len(blocks)):
        for offset in range(3):
            col_letter = get_column_letter(len(index_headers) + 1 + block_no * 3 + offset)
            ws.column_dimensions[col_letter].width = 14
    ws.row_dimensions[band_row].height = 30

    wb.save(output_path_file)
    print(f"Bill Master Report generated: {output_path_file}")
    return output_path_file


def monthly_projectwise_labour_sheet(df_list, output_dir=""):
    """Month × Project billing matrix, one section per (Group, Particulars),
    built from a raw multi-project SRN Report export."""
    df = _parse_srn_export(df_list[0].copy())
    df = df[df['Project Name'].notna()]

    df['SRN Date'] = pd.to_datetime(df['SRN Date'], errors='coerce')
    df = df[df['SRN Date'].notna()]
    df['Month Key'] = df['SRN Date'].dt.to_period('M')

    agg = (
        df.groupby(['Group', 'Particulars', 'Unit', 'Month Key', 'Project Name'], as_index=False)
        .agg({"SRN Qty": "sum", "SRN Amount": "sum"})
    )

    # Sections with the most month-project data cells come first;
    # alphabetical order breaks ties.
    section_sizes = agg.groupby(['Group', 'Particulars', 'Unit']).size()
    sections = sorted(
        section_sizes.index,
        key=lambda k: (-section_sizes[k], str(k[0]).lower(), str(k[1]).lower())
    )

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Monthly Projectwise'

    # ── Styles ───────────────────────────────────────────────────────────────
    title_font = Font(name='Arial', size=16, bold=True, color='1F3864')
    subtitle_font = Font(name='Arial', size=12, bold=True)
    meta_font = Font(name='Arial', size=9, italic=True)
    section_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    section_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    band_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    band_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    sub_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    sub_font = Font(name='Arial', size=9, bold=True, color='1F3864')
    total_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    total_font = Font(name='Arial', size=10, bold=True)
    alt_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
    data_font = Font(name='Arial', size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')
    thin = Side(style='thin', color='B7C4D6')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    NUM_FMT = '#,##0.00'
    sub_headers = ['Billed Qty', 'Billed Rate', 'Billed Amount']

    def write_cell(row, col, value, *, fmt=None, font=data_font, fill=None, align=right):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.font = font
        cell.alignment = align
        cell.border = border
        if fmt:
            cell.number_format = fmt
        if fill:
            cell.fill = fill
        return cell

    # ── Report header ────────────────────────────────────────────────────────
    ws['A1'] = COMPANY_TITLE
    ws['A1'].font = title_font
    ws['A2'] = MONTHLY_PROJECTWISE_LABOUR_REPORT
    ws['A2'].font = subtitle_font
    ws['A3'] = f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A3'].font = meta_font

    current_row = 5
    max_block_count = 1

    for (group, particulars, unit) in sections:
        section_df = agg[
            (agg['Group'] == group) & (agg['Particulars'] == particulars) & (agg['Unit'] == unit)
        ]
        projects = sorted(section_df['Project Name'].unique())
        months = sorted(section_df['Month Key'].unique())
        blocks = projects + (['All Projects'] if len(projects) > 1 else [])
        max_block_count = max(max_block_count, len(blocks))

        lookup = {
            (row['Month Key'], row['Project Name']): (row['SRN Qty'], row['SRN Amount'])
            for _, row in section_df.iterrows()
        }

        # Section title bar across all section columns
        section_end_col = 1 + len(blocks) * 3
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=section_end_col)
        unit_suffix = f"  ({unit})" if unit else ""
        cell = ws.cell(row=current_row, column=1)
        cell.value = f'{group}  —  {particulars}{unit_suffix}'
        cell.fill = section_fill
        cell.font = section_font
        cell.alignment = left
        for c in range(1, section_end_col + 1):
            ws.cell(row=current_row, column=c).border = border
            ws.cell(row=current_row, column=c).fill = section_fill

        # Header band: Month + project blocks
        band_row = current_row + 1
        sub_row = band_row + 1
        ws.merge_cells(start_row=band_row, start_column=1, end_row=sub_row, end_column=1)
        cell = ws.cell(row=band_row, column=1)
        cell.value = 'Month'
        cell.fill = band_fill
        cell.font = band_font
        cell.alignment = center
        ws.cell(row=band_row, column=1).border = border
        ws.cell(row=sub_row, column=1).border = border

        for block_no, project in enumerate(blocks):
            start_col = 2 + block_no * 3
            ws.merge_cells(start_row=band_row, start_column=start_col, end_row=band_row, end_column=start_col + 2)
            cell = ws.cell(row=band_row, column=start_col)
            cell.value = project
            cell.fill = band_fill
            cell.font = band_font
            cell.alignment = center
            for offset, sub in enumerate(sub_headers):
                ws.cell(row=band_row, column=start_col + offset).border = border
                sub_cell = ws.cell(row=sub_row, column=start_col + offset)
                sub_cell.value = sub
                sub_cell.fill = sub_fill
                sub_cell.font = sub_font
                sub_cell.alignment = center
                sub_cell.border = border

        # Month rows
        data_start = sub_row + 1
        col_qty_totals = {p: 0.0 for p in projects}
        col_amt_totals = {p: 0.0 for p in projects}
        for row_no, month in enumerate(months):
            excel_row = data_start + row_no
            row_fill = alt_fill if row_no % 2 == 1 else None
            write_cell(excel_row, 1, month.strftime('%b-%Y'), fill=row_fill, align=left)

            month_qty = 0.0
            month_amt = 0.0
            for block_no, project in enumerate(projects):
                start_col = 2 + block_no * 3
                if (month, project) in lookup:
                    qty, amt = lookup[(month, project)]
                    qty, amt = float(qty), float(amt)
                    rate = (amt / qty) if qty else 0.0
                    write_cell(excel_row, start_col, round(qty, 2), fmt=NUM_FMT, fill=row_fill)
                    write_cell(excel_row, start_col + 1, round(rate, 2), fmt=NUM_FMT, fill=row_fill)
                    write_cell(excel_row, start_col + 2, round(amt, 2), fmt=NUM_FMT, fill=row_fill)
                    month_qty += qty
                    month_amt += amt
                    col_qty_totals[project] += qty
                    col_amt_totals[project] += amt
                else:
                    for offset in range(3):
                        write_cell(excel_row, start_col + offset, None, fill=row_fill)

            if len(projects) > 1:
                start_col = 2 + len(projects) * 3
                rate = (month_amt / month_qty) if month_qty else 0.0
                write_cell(excel_row, start_col, round(month_qty, 2), fmt=NUM_FMT, font=total_font, fill=row_fill)
                write_cell(excel_row, start_col + 1, round(rate, 2), fmt=NUM_FMT, font=total_font, fill=row_fill)
                write_cell(excel_row, start_col + 2, round(month_amt, 2), fmt=NUM_FMT, font=total_font, fill=row_fill)

        # Section total row
        total_row = data_start + len(months)
        write_cell(total_row, 1, 'Total', font=total_font, fill=total_fill, align=left)
        all_qty = 0.0
        all_amt = 0.0
        for block_no, project in enumerate(projects):
            start_col = 2 + block_no * 3
            qty = col_qty_totals[project]
            amt = col_amt_totals[project]
            rate = (amt / qty) if qty else 0.0
            write_cell(total_row, start_col, round(qty, 2), fmt=NUM_FMT, font=total_font, fill=total_fill)
            write_cell(total_row, start_col + 1, round(rate, 2), fmt=NUM_FMT, font=total_font, fill=total_fill)
            write_cell(total_row, start_col + 2, round(amt, 2), fmt=NUM_FMT, font=total_font, fill=total_fill)
            all_qty += qty
            all_amt += amt
        if len(projects) > 1:
            start_col = 2 + len(projects) * 3
            rate = (all_amt / all_qty) if all_qty else 0.0
            write_cell(total_row, start_col, round(all_qty, 2), fmt=NUM_FMT, font=total_font, fill=total_fill)
            write_cell(total_row, start_col + 1, round(rate, 2), fmt=NUM_FMT, font=total_font, fill=total_fill)
            write_cell(total_row, start_col + 2, round(all_amt, 2), fmt=NUM_FMT, font=total_font, fill=total_fill)

        # Two blank rows between sections
        current_row = total_row + 3

    # ── Layout polish ─────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 14
    for block_no in range(max_block_count):
        for offset in range(3):
            col_letter = get_column_letter(2 + block_no * 3 + offset)
            ws.column_dimensions[col_letter].width = 14
    ws.freeze_panes = 'A5'

    output_path_file = get_output_excel("All Projects", "Monthly_Projectwise_Labour_Report", ".xlsx", output_dir)
    wb.save(output_path_file)
    print(f"Monthly Projectwise Labour Report generated: {output_path_file}")
    return output_path_file
